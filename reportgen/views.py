from django.shortcuts import render
from logger.models import CallLog
from datetime import datetime, timedelta
from openpyxl import Workbook
from django.http import HttpResponse
from tempfile import NamedTemporaryFile

rate = 12.0

# Create your views here.
def index(request):
    report_range = request.GET.get("range")

    if report_range.strip() == "":
        report_range = "weekly"

    report_range = report_range.lower()
    since = None
    if report_range == "weekly":
        # get time of last sunday 12am
        today = datetime.today()
        since = today - timedelta(days=today.weekday())
    elif report_range == "monthly":
        # get the timestamp of the first day of the month
        today = datetime.today()
        since = datetime(today.year, today.month, 1)

    # get all call logs since 'since'
    call_logs = CallLog.objects.filter(time_booked__gte=since)

    wb = Workbook()
    ws = wb.active  # active sheet

    # insert headers
    ws.append(["Date", "Description", "Amount", "Currency"])

    i, j = 2, 1
    # walk through all call logs, and add them to the sheet
    for call_log in call_logs:
        ws.cell(row=i, column=j).value = date_of(call_log.time_booked)

        description = f"{call_time_of(call_log)} {''.join(call_log.destination.split(' '))} {call_log.fullname}"

        ws.cell(row=i, column=j + 1).value = description

        ws.cell(row=i, column=j + 2).value = call_time_of(call_log) * rate
        ws.cell(row=i, column=j + 3).value = "NGN"
        i += 1

    # send the file to the browser
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=report.xlsx"
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response.write(tmp.read())
    return response


def date_of(timestamp):
    return timestamp.strftime("%d-%m-%y")


def call_time_of(log):
    if log.remark.lower() != "success":
        return 0
    time_conn = log.time_connected
    time_fin = log.time_finished
    # return the difference between the two times in minutes
    return (time_fin - time_conn).seconds // 60
